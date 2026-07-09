import os
import json
import datetime
import calendar
import webbrowser
import threading
import hashlib
from flask import Flask, request, jsonify, render_template, session

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = 'workmap_secure_session_key_secret_987654321'

DB_FILE = 'schedule_db.json'
USERS_FILE = 'users_db.json'
EXCEL_FILE = 'Azimoff.xlsx'

def format_time(t):
    if t is None:
        return ""
    if isinstance(t, datetime.time):
        return t.strftime("%H:%M")
    if isinstance(t, str):
        return t
    return ""

def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def get_usd_uzs_rate():
    # Try Central Bank of Uzbekistan first
    try:
        import urllib.request
        req = urllib.request.Request(
            'https://cbu.uz/ru/arkhiv-kursov-valyut/json/',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            for item in data:
                if item.get('Ccy') == 'USD':
                    rate = float(item.get('Rate', 12050.0))
                    if rate > 0:
                        return rate
    except Exception as e:
        print(f"CBU rate fetch failed: {e}")
        
    # Try backup Open Exchange Rate API
    try:
        import urllib.request
        with urllib.request.urlopen('https://open.er-api.com/v6/latest/USD', timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            rates = data.get('rates', {})
            if 'UZS' in rates:
                rate = float(rates['UZS'])
                if rate > 0:
                    return rate
    except Exception as e:
        print(f"Backup rate fetch failed: {e}")
        
    return 12050.0 # fallback rate

def read_users():
    if not os.path.exists(USERS_FILE):
        return {}
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def write_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=2, ensure_ascii=False)

def read_db():
    if not os.path.exists(DB_FILE):
        return {}
    try:
        with open(DB_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def write_db(data):
    with open(DB_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def get_user_data(username):
    db = read_db()
    return db.get(username, {})

def write_user_data(username, data):
    db = read_db()
    db[username] = data
    write_db(db)

def migrate_db_format():
    if not os.path.exists(DB_FILE):
        return
    try:
        with open(DB_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Check if the database has YYYY-MM keys at root level (old format)
        has_old_keys = any(len(k) == 7 and '-' in k for k in data.keys())
        
        if has_old_keys:
            print("Converting database to user-nested format...")
            new_data = {
                "azimoff": data
            }
            write_db(new_data)
            
            # Create default user 'azimoff' with password 'azimoff'
            users = read_users()
            if "azimoff" not in users:
                users["azimoff"] = hash_password("azimoff")
                write_users(users)
            print("Migration complete! Default data nested under user 'azimoff' with password 'azimoff'.")
    except Exception as e:
        print(f"Error migrating database format: {e}")

def migrate_users_format():
    try:
        users = read_users()
        changed = False
        for u, val in list(users.items()):
            if isinstance(val, str):
                users[u] = {
                    "password": val,
                    "avatar": None
                }
                changed = True
        if changed:
            write_users(users)
            print("Users database format successfully migrated to dictionary layout.")
    except Exception as e:
        print(f"Error migrating users format: {e}")

def migrate_excel():
    # If users or database exists, skip Excel migration
    if os.path.exists(DB_FILE) or os.path.exists(USERS_FILE):
        return
    
    if not os.path.exists(EXCEL_FILE):
        return

    print("Migrating Azimoff.xlsx...")
    db = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        # Import 'Aprel N' as '2024-04'
        if 'Aprel N' in wb.sheetnames:
            sheet = wb['Aprel N']
            days = {}
            for r in range(5, 36):
                day_val = sheet.cell(r, 1).value
                if day_val is None:
                    continue
                day_num = int(float(day_val))
                keldi = format_time(sheet.cell(r, 3).value)
                ketdi = format_time(sheet.cell(r, 4).value)
                plan = sheet.cell(r, 5).value or 0
                days[str(day_num)] = {
                    "keldi": keldi,
                    "ketdi": ketdi,
                    "plan": plan
                }
            
            salary = sheet.cell(43, 4).value or 700.0
            bonus = sheet.cell(44, 4).value or 300.0
            kurs = sheet.cell(51, 4).value or 12050.0
            avans = sheet.cell(48, 4).value or 1200000.0
            
            db["2024-04"] = {
                "days": days,
                "salary": salary,
                "bonus": bonus,
                "kurs": kurs,
                "avans": avans,
                "calc_method": "proportional"
            }
            print("Successfully migrated 'Aprel N' to '2024-04'")
            
    except Exception as e:
        print(f"Error during Excel migration: {e}")
    
    # Store migrated data under default user 'azimoff'
    if db:
        write_user_data("azimoff", db)
        users = read_users()
        users["azimoff"] = hash_password("azimoff")
        write_users(users)

def get_default_month_data(year_month, username):
    current_live_rate = get_usd_uzs_rate()
    try:
        year, month = map(int, year_month.split('-'))
    except:
        return {"days": {}, "salary": 0.0, "bonus": 0.0, "kurs": current_live_rate, "avans": 0.0, "calc_method": "proportional"}
        
    days = {}
    _, num_days = calendar.monthrange(year, month)
    
    for day in range(1, num_days + 1):
        date_obj = datetime.date(year, month, day)
        wday = date_obj.weekday() # 0-6 (0=Monday)
        
        # Mon-Fri: 9h, Sat: 8h (since Saturday is short), Sun: 0h
        if wday < 5:
            plan = 9.0
        elif wday == 5:
            plan = 8.0 # Saturday plan
        else:
            plan = 0.0
            
        days[str(day)] = {
            "keldi": "",
            "ketdi": "",
            "plan": plan
        }
        
    # Inherit latest settings for this user
    user_data = get_user_data(username)
    salary = 0.0
    bonus = 0.0
    unplanned_bonus = 0.0
    kurs = current_live_rate
    avans = 0.0
    card_transfer = 0.0
    
    if user_data:
        valid_keys = [k for k in user_data.keys() if len(k) == 7 and '-' in k]
        if valid_keys:
            latest_key = sorted(valid_keys)[-1]
            salary = user_data[latest_key].get("salary", salary)
            bonus = user_data[latest_key].get("bonus", bonus)
            unplanned_bonus = user_data[latest_key].get("unplanned_bonus", unplanned_bonus)
            kurs = user_data[latest_key].get("kurs", kurs)
            avans = user_data[latest_key].get("avans", avans)
            card_transfer = user_data[latest_key].get("card_transfer", card_transfer)
            
    return {
        "days": days,
        "salary": salary,
        "bonus": bonus,
        "unplanned_bonus": unplanned_bonus,
        "kurs": kurs,
        "avans": avans,
        "card_transfer": card_transfer,
        "calc_method": "proportional"
    }

# Flask Endpoints
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/user', methods=['GET'])
def get_current_user():
    username = session.get('username')
    if not username:
        return jsonify({"username": None}), 401
    users = read_users()
    user_val = users.get(username)
    avatar = None
    verifix_url = "https://app.verifix.com"
    verifix_login = ""
    verifix_has_password = False
    if isinstance(user_val, dict):
        avatar = user_val.get('avatar')
        verifix_url = user_val.get('verifix_url', verifix_url) or "https://app.verifix.com"
        verifix_login = user_val.get('verifix_login', '')
        verifix_has_password = bool(user_val.get('verifix_password'))
    return jsonify({
        "username": username,
        "avatar": avatar,
        "verifix_url": verifix_url,
        "verifix_login": verifix_login,
        "verifix_has_password": verifix_has_password
    })

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({"success": False, "message": "Введите имя пользователя и пароль"}), 400
        
    users = read_users()
    hashed = hash_password(password)
    
    if username in users:
        user_val = users[username]
        stored_hash = user_val["password"] if isinstance(user_val, dict) else user_val
        if stored_hash == hashed:
            session['username'] = username
            avatar = user_val.get("avatar") if isinstance(user_val, dict) else None
            return jsonify({"success": True, "avatar": avatar})
        else:
            return jsonify({"success": False, "message": "Неверный пароль"}), 401
    else:
        # Register new user automatically
        users[username] = {
            "password": hashed,
            "avatar": None
        }
        write_users(users)
        session['username'] = username
        return jsonify({"success": True, "registered": True, "avatar": None})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('username', None)
    return jsonify({"success": True})

@app.route('/api/update-profile', methods=['POST'])
def update_profile():
    current_user = session.get('username')
    if not current_user:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
        
    data = request.get_json() or {}
    new_username = data.get('new_username', '').strip().lower()
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    confirm_password = data.get('confirm_password', '')
    
    users = read_users()
    db = read_db()
    
    # Check avatar change
    avatar_sent = 'avatar' in data
    avatar_changed = False
    if avatar_sent:
        user_val = users.get(current_user)
        existing_avatar = user_val.get('avatar') if isinstance(user_val, dict) else None
        if data.get('avatar') != existing_avatar:
            avatar_changed = True
            
    username_changed = False
    if new_username and new_username != current_user:
        if new_username in users:
            return jsonify({"success": False, "message": "Это имя пользователя уже занято"}), 400
        username_changed = True
        
    password_changed = False
    if new_password:
        if new_password != confirm_password:
            return jsonify({"success": False, "message": "Новые пароли не совпадают"}), 400
        if len(new_password) < 4:
            return jsonify({"success": False, "message": "Пароль должен быть не менее 4 символов"}), 400
        password_changed = True
        
    verifix_url_sent = 'verifix_url' in data
    verifix_login_sent = 'verifix_login' in data
    verifix_password_sent = 'verifix_password' in data
    
    something_changed = (username_changed or password_changed or avatar_changed or 
                         verifix_url_sent or verifix_login_sent or verifix_password_sent)
    
    # 1. Verification of current password
    credentials_changed = username_changed or password_changed
    if credentials_changed:
        if not current_password:
            return jsonify({"success": False, "message": "Введите текущий пароль для подтверждения изменений"}), 400
        user_val = users.get(current_user)
        stored_hash = user_val["password"] if isinstance(user_val, dict) else user_val
        if stored_hash != hash_password(current_password):
            return jsonify({"success": False, "message": "Неверный текущий пароль"}), 400
            
    # Apply changes
    target_username = current_user
    user_info = users.get(current_user)
    if not isinstance(user_info, dict):
        user_info = {
            "password": user_info,
            "avatar": None
        }
        users[target_username] = user_info
        
    if username_changed:
        target_username = new_username
        # Copy to new username key
        users[new_username] = user_info
        if current_user in users:
            del users[current_user]
        
        # Copy schedule to new key
        if current_user in db:
            db[new_username] = db[current_user]
            del db[current_user]
            
    if password_changed:
        users[target_username]["password"] = hash_password(new_password)
        
    if avatar_changed:
        users[target_username]["avatar"] = data.get('avatar')
        
    if verifix_url_sent:
        users[target_username]["verifix_url"] = data.get('verifix_url', '').strip()
    if verifix_login_sent:
        users[target_username]["verifix_login"] = data.get('verifix_login', '').strip()
    if verifix_password_sent:
        users[target_username]["verifix_password"] = data.get('verifix_password', '')
        
    # Save databases if anything changed
    if something_changed:
        write_users(users)
        if username_changed:
            write_db(db)
        session['username'] = target_username
        
    # Get final avatar
    final_avatar = users[target_username].get("avatar") if isinstance(users[target_username], dict) else None
        
    return jsonify({"success": True, "username": target_username, "avatar": final_avatar})

@app.route('/api/current-usd-rate', methods=['GET'])
def api_current_usd_rate():
    rate = get_usd_uzs_rate()
    return jsonify({"rate": rate})

@app.route('/api/schedule', methods=['GET'])
def get_schedule():
    username = session.get('username')
    if not username:
        return jsonify({"error": "Unauthorized"}), 401
        
    year_month = request.args.get('year_month')
    if not year_month:
        return jsonify({"error": "Missing year_month parameter"}), 400
        
    user_data = get_user_data(username)
    if year_month in user_data:
        month_data = user_data[year_month]
        if "unplanned_bonus" not in month_data:
            month_data["unplanned_bonus"] = 0.0
        if "card_transfer" not in month_data:
            month_data["card_transfer"] = 0.0
        return jsonify(month_data)
    else:
        return jsonify(get_default_month_data(year_month, username))

@app.route('/api/schedule', methods=['POST'])
def save_schedule():
    username = session.get('username')
    if not username:
        return jsonify({"error": "Unauthorized"}), 401
        
    data = request.get_json()
    if not data or 'year_month' not in data:
        return jsonify({"error": "Invalid request data"}), 400
        
    year_month = data['year_month']
    user_data = get_user_data(username)
    
    user_data[year_month] = {
        "days": data.get("days", {}),
        "salary": float(data.get("salary", 700.0)),
        "bonus": float(data.get("bonus", 0.0)),
        "unplanned_bonus": float(data.get("unplanned_bonus", 0.0)),
        "kurs": float(data.get("kurs", 12050.0)),
        "avans": float(data.get("avans", 0.0)),
        "card_transfer": float(data.get("card_transfer", 0.0)),
        "calc_method": "proportional"
    }
    write_user_data(username, user_data)
    return jsonify({"success": True})

@app.route('/api/history', methods=['GET'])
def get_history():
    username = session.get('username')
    if not username:
        return jsonify({"error": "Unauthorized"}), 401
        
    user_data = get_user_data(username)
    history = []
    
    keys = [k for k in user_data.keys() if len(k) == 7 and '-' in k]
    for key in sorted(keys):
        month_data = user_data[key]
        days = month_data.get("days", {})
        
        total_plan = sum(day.get("plan", 0.0) for day in days.values())
        total_actual = 0.0
        for day in days.values():
            keldi = day.get("keldi", "")
            ketdi = day.get("ketdi", "")
            if keldi and ketdi:
                try:
                    a_h, a_m = map(int, keldi.split(':'))
                    d_h, d_m = map(int, ketdi.split(':'))
                    total_actual += (d_h * 60 + d_m - (a_h * 60 + a_m)) / 60.0
                except:
                    pass
                    
        salary = float(month_data.get("salary", 700.0))
        bonus = float(month_data.get("bonus", 0.0))
        unplanned_bonus = float(month_data.get("unplanned_bonus", 0.0))
        kurs = float(month_data.get("kurs", 12050.0))
        avans = float(month_data.get("avans", 0.0))
        card_transfer = float(month_data.get("card_transfer", 0.0))
        
        final_usd = 0.0
        if total_plan > 0:
            final_usd = (salary + bonus + unplanned_bonus) / total_plan * total_actual
            
        total_uzs = final_usd * kurs
        remaining_cash = total_uzs - avans - card_transfer
        
        history.append({
            "year_month": key,
            "total_plan": round(total_plan, 2),
            "total_actual": round(total_actual, 2),
            "total_usd": round(final_usd, 2),
            "total_uzs": round(total_uzs),
            "remaining_cash": round(remaining_cash)
        })
        
    return jsonify(history)

def extract_time_records_from_json(data):
    import re
    date_pat = re.compile(r'^(\d{2})\.(\d{2})\.(\d{4})$') # DD.MM.YYYY
    date_iso_pat = re.compile(r'^(\d{4})-(\d{2})-(\d{2})$') # YYYY-MM-DD
    time_pat = re.compile(r'^(\d{2}):(\d{2})$') # HH:MM
    time_sec_pat = re.compile(r'^(\d{2}):(\d{2}):(\d{2})$') # HH:MM:SS
    
    records = []
    
    def walk(node):
        if isinstance(node, list):
            for item in node:
                walk(item)
        elif isinstance(node, dict):
            found_date = None
            found_times = []
            for k, v in node.items():
                if isinstance(v, str):
                    v = v.strip()
                    if date_pat.match(v) or date_iso_pat.match(v):
                        found_date = v
                    elif time_pat.match(v) or time_sec_pat.match(v):
                        found_times.append(v[:5])
            
            if found_date and found_times:
                for t in found_times:
                    records.append((found_date, t))
            else:
                for v in node.values():
                    walk(v)
                    
    walk(data)
    return records

def run_verifix_sync(verifix_url, verifix_login, verifix_password, year_month):
    import urllib.request
    import urllib.parse
    import json
    import hashlib
    import re
    import datetime
    from http.cookiejar import CookieJar
    
    debug_logs = []
    def log_debug(msg):
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        debug_logs.append(f"[{timestamp}] {msg}")
        print(f"[VERIFIX_SYNC] {msg}")

    def save_debug_file():
        try:
            with open('verifix_debug.log', 'w', encoding='utf-8') as f:
                f.write("\n".join(debug_logs))
        except Exception as ex:
            print(f"Failed writing debug log: {ex}")

    log_debug(f"Starting Verifix Sync. URL: {verifix_url}, Login: {verifix_login}, Month: {year_month}")
    
    try:
        year, month = map(int, year_month.split('-'))
    except:
        log_debug("Invalid month format")
        save_debug_file()
        return None, "Неверный формат месяца/года"
        
    _, num_days = calendar.monthrange(year, month)
    begin_date = f"01.{month:02d}.{year}"
    end_date = f"{num_days:02d}.{month:02d}.{year}"
    
    log_debug(f"Date range: {begin_date} to {end_date}")
    
    sha1_pwd = hashlib.sha1(verifix_password.encode('utf-8')).hexdigest()
    
    cj = CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    
    login_url = f"{verifix_url.rstrip('/')}/b/biruni/s$log_in"
    login_data = urllib.parse.urlencode({
        'login': verifix_login,
        'password': sha1_pwd,
        'lang_code': 'ru'
    }).encode('utf-8')
    
    req = urllib.request.Request(
        login_url,
        data=login_data,
        headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'X-Requested-With': 'XMLHttpRequest'
        }
    )
    
    log_debug(f"Sending login request to: {login_url}")
    try:
        with opener.open(req, timeout=10) as resp:
            resp_body = resp.read().decode('utf-8')
            try:
                res_json = json.loads(resp_body)
            except Exception as json_err:
                log_debug(f"Failed to parse login response JSON: {json_err}")
                save_debug_file()
                return None, f"Ответ сервера Verifix не является JSON. Ответ: {resp_body[:200]}"
                
            if res_json.get('status') != 'logged_in':
                err_msg = res_json.get('error') or res_json.get('message') or "Неверный логин или пароль"
                log_debug(f"Login rejected: {err_msg}")
                save_debug_file()
                return None, f"Ошибка авторизации Verifix: {err_msg}"
            
            log_debug("Login successful!")
    except urllib.error.HTTPError as he:
        try:
            err_body = he.read().decode('utf-8')
            log_debug(f"Login HTTP Error {he.code}: {err_body[:1000]}")
            err_json = json.loads(err_body)
            err_msg = err_json.get('error') or err_json.get('message') or str(he)
        except:
            err_msg = str(he)
        save_debug_file()
        return None, f"Ошибка HTTP при авторизации: {err_msg}"
    except Exception as e:
        log_debug(f"Login failed connection exception: {e}")
        save_debug_file()
        return None, f"Не удалось подключиться к Verifix: {str(e)}"
        
    updated_days = {}
    
    log_debug("Querying daily dashboard table data...")
    for day in range(1, num_days + 1):
        date_str = f"{day:02d}.{month:02d}.{year}"
        
        payload = {
            "p": {
                "column": ["name", "input_time", "output_time", "begin_time", "kind_name"],
                "filter": [],
                "sort": ["name"],
                "offset": 0,
                "limit": 50
            },
            "d": {
                "division_id": [],
                "job_id": [],
                "schedule_id": [],
                "rank_id": [],
                "location_id": [],
                "fte_id": [],
                "date": date_str
            }
        }
        
        url = f"{verifix_url.rstrip('/')}/b/vhr/intro/dashboard:table"
        req_data = json.dumps(payload).encode('utf-8')
        
        req = urllib.request.Request(
            url,
            data=req_data,
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Content-Type': 'application/json; charset=UTF-8',
                'X-Requested-With': 'XMLHttpRequest'
            }
        )
        
        try:
            with opener.open(req, timeout=10) as resp:
                resp_body = resp.read().decode('utf-8')
                res_json = json.loads(resp_body)
                
                rows = res_json.get("data", [])
                if not rows:
                    log_debug(f"No records for {date_str}")
                    continue
                    
                # Find matching row for user
                matched_row = None
                if len(rows) == 1:
                    matched_row = rows[0]
                else:
                    # Filter by name similarity
                    login_parts = re.split(r'[^a-zA-Z0-9]', verifix_login.split('@')[0].lower())
                    login_parts = [p for p in login_parts if len(p) > 2]
                    for row in rows:
                        row_name = row[0].lower()
                        if all(p in row_name for p in login_parts):
                            matched_row = row
                            break
                            
                if matched_row:
                    name, input_time, output_time, begin_time, kind_name = matched_row
                    log_debug(f"Matched row for {date_str}: {name}, in: {input_time}, out: {output_time}")
                    
                    keldi = input_time if input_time else ""
                    ketdi = output_time if output_time else ""
                    
                    if keldi or ketdi:
                        updated_days[str(day)] = {
                            "keldi": keldi,
                            "ketdi": ketdi
                        }
                else:
                    log_debug(f"Could not match user row on {date_str} from {len(rows)} rows.")
        except Exception as e:
            log_debug(f"Error querying dashboard table on {date_str}: {e}")
            
    save_debug_file()
    if not updated_days:
        return None, f"Авторизовано успешно, но не удалось найти данные посещаемости за период {begin_date} - {end_date} в Verifix."
        
    return updated_days, "Успешно извлечено через форму dashboard:table"

@app.route('/api/verifix/sync', methods=['POST'])
def verifix_sync():
    username = session.get('username')
    if not username:
        return jsonify({"error": "Unauthorized"}), 401
        
    data = request.get_json() or {}
    year_month = data.get('year_month')
    if not year_month:
        return jsonify({"error": "Missing year_month parameter"}), 400
        
    users = read_users()
    user_val = users.get(username)
    if not isinstance(user_val, dict):
        return jsonify({"error": "User profile not configured"}), 400
        
    verifix_url = user_val.get('verifix_url', 'https://app.verifix.com').strip()
    verifix_login = user_val.get('verifix_login', '').strip()
    verifix_password = user_val.get('verifix_password', '')
    
    if not verifix_login or not verifix_password:
        return jsonify({"success": False, "message": "Настройки Verifix не заполнены в вашем профиле!"}), 400
        
    try:
        updated_days, message = run_verifix_sync(verifix_url, verifix_login, verifix_password, year_month)
        if updated_days is None:
            return jsonify({"success": False, "message": message})
            
        user_data = get_user_data(username)
        if year_month not in user_data:
            user_data[year_month] = get_default_month_data(year_month, username)
            
        month_data = user_data[year_month]
        days = month_data.get("days", {})
        
        count = 0
        for d_str, times in updated_days.items():
            if d_str in days:
                days[d_str]["keldi"] = times["keldi"]
                days[d_str]["ketdi"] = times["ketdi"]
                count += 1
                
        month_data["days"] = days
        user_data[year_month] = month_data
        write_user_data(username, user_data)
        
        return jsonify({
            "success": True, 
            "message": f"Успешно импортировано из Verifix! Обновлено дней: {count}. ({message})",
            "days": days
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Ошибка синхронизации: {str(e)}"}), 500

def open_browser():
    webbrowser.open_new("http://localhost:5000")

if __name__ == '__main__':
    migrate_db_format()
    migrate_users_format()
    migrate_excel()
    
    if os.environ.get("WERKZEUG_RUN_MAIN") != "true" and os.environ.get("FLASK_ENV") != "production":
        threading.Timer(1.2, open_browser).start()
        
    # Bind to 0.0.0.0 for deploying
    app.run(host='0.0.0.0', port=5000, debug=True)
